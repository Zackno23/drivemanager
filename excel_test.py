"""
送られてきたデータからエクセルデータを出力
実際に出力するのは
"""
import openpyxl
import sqlite3
from datetime import datetime


# sqlからデータを読み出し、緯度のリスト、軽度のリストを作成
def sql_to_excel():
    conn = sqlite3.connect('gps_log.sqlite')
    cur = conn.cursor()

    cur.execute('select * from log')
    gps_list = cur.fetchall()
    gps_long = [i[0] for i in gps_list]
    gps_lat = [i[1] for i in gps_list]
    cur.close()
    conn.close()
    return (gps_long, gps_lat)


if __name__ == '__main__':
    gps_long = sql_to_excel()[0]
    gps_lat = sql_to_excel()[1]

    """
    Excelファイル(xlsx)を作るだけ
    `example.xlsx` が実行したカレントにできる
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    list = str(datetime.today()).split('-')
    list[2] = list[2][:2]
    date = f'{list[0]}年{list[1]}月{list[2]}日'
    ws.title = date

    ws.cell(row=2, column=2, value=f"{date}の運行管理")
    # row=3, column=3 を起点として、gps_lonのデータを入れていく
    for i in range(len(gps_lat)):
        ws.cell(row=i + 3, column=2, value=gps_long[i])
        ws.cell(row=i + 3, column=3, value=gps_lat[i])
    ws.cell(row=5, column=1, value="")
    # 保存
    wb.save('./運行管理表.xlsx')
