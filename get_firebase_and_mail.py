from firebase import firebase
import datetime
import openpyxl
from send_gmail import sendGmailAttach
from openpyxl.styles.borders import Border, Side

firebase = firebase.FirebaseApplication('https://drive-manager-18863.firebaseio.com/', None)
all_firevase_data = firebase.get('/drive-manager-18863/', '')

"""
Excelファイル(xlsx)を作るだけ
`example.xlsx` が実行したカレントディレクトリにできる
"""

date_list = str(datetime.datetime.today()).split('-')
date_list[2] = date_list[2][:2]
date = f'{date_list[0]}年{date_list[1]}月{date_list[2]}日'

wb = openpyxl.load_workbook('./運行管理表.xlsx')

if date[:8] in wb.sheetnames:
    ws = wb[date[:8]]
else:
    ws = wb.create_sheet(date[:8])
ws.title = date[:8]
ws.merge_cells('B2:D2')

with open('koutei.txt', 'r') as f:
    koutei = int(f.read())

total_distance = 0
for i in range(koutei + 1):
    listlist = list(dict(all_firevase_data).values())[0]
    total_distance += list(list(dict(all_firevase_data).values())[0][f'行程{i}']['目的地：市役所']['内容：MTG'].values())[1]["合計距離"]

goal = '市役所'
work = 'MTG'
values = []
maxRow = ws.max_row
for cell in ws['B']:
    values.append(cell.value)

side = Side(style='thin', color='000000')
border = Border(top=side, bottom=side, left=side, right=side)


if f"{date}の運行管理" in values:
    for i in range(koutei + 1):
        ws.cell(row=maxRow + 1 + i, column=2, value=float(total_distance))
        ws.cell(row=maxRow + 1 + i, column=3, value=goal)
        ws.cell(row=maxRow + 1 + i, column=4, value=work)
    for row in ws:
        for cell in row:
            if ws[cell.coordinate].value:
                ws[cell.coordinate].border = border

else:

    ws.cell(row=maxRow + 2, column=2, value=f"{date}の運行管理")
    ws.cell(row=maxRow + 3, column=2, value="走行距離")
    ws.cell(row=maxRow + 3, column=3, value="目的地")
    ws.cell(row=maxRow + 3, column=4, value="活動内容")
    side = Side(style='thin', color='000000')
    border = Border(top=side, bottom=side, left=side, right=side)
    for i in range(koutei + 1):
        ws.cell(row=maxRow + 4 + i, column=2, value=float(total_distance))
        ws.cell(row=maxRow + 4 + i, column=3, value=goal)
        ws.cell(row=maxRow + 4 + i, column=4, value=work)
    for row in ws:
        for cell in row:
            if ws[cell.coordinate].value:
                ws[cell.coordinate].border = border
maxRow = ws.max_row

print(values)
sum_list = []
sum_list.append(values[-1])
print(sum_list)
counter = 0
for items in values:
    if counter < len(values)-1:
        if items is not None and values[counter + 1] is None:
            sum_list.append(items)
    counter += 1
ws.cell(row=5, column=6, value=sum(sum_list))
print(sum_list)





wb.save('./運行管理表.xlsx')

sendGmailAttach(f'{date}の運行管理表', f'{date}の運行管理表です。ご査収ください。\n\n吉田  力')
print("sent")
