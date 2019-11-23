from firebase import firebase
import datetime
import openpyxl
from send_gmail import sendGmailAttach

firebase = firebase.FirebaseApplication('https://drive-manager-18863.firebaseio.com/', None)
all_firevase_data = firebase.get('/drive-manager-18863/', '')

"""
Excelファイル(xlsx)を作るだけ
`example.xlsx` が実行したカレントにできる
"""
wb = openpyxl.Workbook()
ws = wb.active
date_list = str(datetime.datetime.today()).split('-')
date_list[2] = date_list[2][:2]
date = f'{date_list[0]}年{date_list[1]}月{date_list[2]}日'
ws.title = date
ws.cell(row=2, column=2, value=f"{date}の運行管理")


with open('koutei.txt','r') as f:
    koutei = int(f.read())
total_distance = 0
for i in range(koutei+1):
    total_distance = str(all_firevase_data['距離0']['合計'])
    gaol = list(dict(all_firevase_data[str(datetime.datetime.today())[:11]][f'行程{koutei}']).keys())[0][4:]
    # row=3, column=3 を起点として、gps_lonのデータを入れていく
    ws.cell(row=3 + i, column=2, value=total_distance)
    ws.cell(row=3 + i, column=3, value=gaol)
    # 保存
    wb.save('./運行管理表.xlsx')


sendGmailAttach(f'{date}の運行管理表',f'{date}の運行管理表です。ご査収ください。\n\n吉田  力')
print("sent")









# def sendGmailAttach():
#     sender, password = "chikara.f.yoshida@gmail.com", "Zackno213" # 送信元メールアドレスとgmailへのログイン情報
#     to = 'hima.hima.0427@gmail.com'  # 送信先メールアドレス
#     sub = f'{datetime.datetime.now()} 運行管理' #メール件名
#     body = '画像添付してるよ'  # メール本文
#     host, port = 'smtp.gmail.com', 587
#
#     # メールヘッダー
#     msg = MIMEMultipart()
#     msg['Subject'] = sub
#     msg['From'] = sender
#     msg['To'] = to
#
#     # メール本文
#     body = MIMEText(body)
#     msg.attach(body)
#
#     # 添付ファイルの設定
#     attach_file = {'name': '運行管理表.xlsx', 'path': '../運行管理表.xlsx'} # nameは添付ファイル名。pathは添付ファイルの位置を指定
#     attachment = MIMEBase('application', 'octet-stream')
#     file = open(attach_file['path'], 'rb+')
#     attachment.set_payload(file.read())
#     file.close()
#     encoders.encode_base64(attachment)
#     attachment.add_header("Content-Disposition", "attachment", filename=attach_file['name'])
#     msg.attach(attachment)
#
#     # gmailへ接続(SMTPサーバーとして使用)
#     gmail=SMTP("smtp.gmail.com", 587)
#     gmail.starttls() # SMTP通信のコマンドを暗号化し、サーバーアクセスの認証を通す
#     gmail.login(sender, password)
#     gmail.send_message(msg)
#
# if __name__ == '__main__':
#     sendGmailAttach()
#     print('メールが送信されました')
